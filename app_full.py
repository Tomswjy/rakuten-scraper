"""
乐天商品爬虫 Web 应用 - 完整版
直接使用原始爬虫代码的全部功能
"""
import os
import io
import re
import sys
import time
import random
import threading
import uuid
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, Response
from flask_cors import CORS
import pandas as pd

# 导入原始爬虫模块的所有函数
from scraper import (
    proxies, PROXY_PORT,
    safe_translate,
    extract_price,
    resolve_pr_link,
    extract_dynamic_specs,
    check_has_video,
    extract_selling_points,
    analyze_selling_points_ai,
    extract_product_features,
    get_categories_from_page,
    get_category_ranking,
    extract_ranking_info,
    analyze_reviews,
    get_product_details,
    USE_AI_FEATURES,
    ENABLE_TRANSLATION,
)
from curl_cffi import requests as cffi_requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# 任务存储
tasks = {}

# 实时日志存储
task_logs = {}


def log_message(task_id, message):
    """记录任务日志"""
    if task_id not in task_logs:
        task_logs[task_id] = []
    timestamp = datetime.now().strftime("%H:%M:%S")
    task_logs[task_id].append(f"[{timestamp}] {message}")
    print(f"[{task_id[:8]}] {message}")


def run_scraper_full(task_id, keyword, pages=1, deep_limit=5, enable_ai=True, enable_translate=True):
    """执行完整版爬虫任务"""
    task = tasks[task_id]
    task['status'] = 'running'
    task['progress'] = 0
    task['message'] = f'开始搜索关键词: {keyword}'
    log_message(task_id, f"🚀 启动爬虫 | 关键词: {keyword} | 页数: {pages} | 深度: {deep_limit}")

    base_url = "https://search.rakuten.co.jp/search/mall/{}/"
    raw_data = []

    try:
        # ==================== Step 1: 列表抓取 ====================
        for page in range(1, pages + 1):
            url = base_url.format(keyword) + f"?p={page}"
            task['message'] = f'正在抓取第 {page}/{pages} 页列表...'
            log_message(task_id, f"📡 抓取列表页 {page}/{pages}: {url}")

            res = cffi_requests.get(url, impersonate="chrome120", timeout=30, proxies=proxies)
            soup = BeautifulSoup(res.text, 'html.parser')

            # 市场饱和度
            page_text = soup.get_text()
            total_txt = "未知"
            count_match = re.search(r'[\(（]([\d,]+)件[\)）]', page_text)
            if count_match and len(count_match.group(1).replace(',', '')) >= 3:
                total_txt = count_match.group(1)

            if total_txt == "未知":
                count_match = re.search(r'([\d,]{5,})件', page_text)
                if count_match:
                    total_txt = count_match.group(1)

            log_message(task_id, f"   📊 搜索结果总数: {total_txt}件")

            items = soup.select('.searchresultitem')
            if not items:
                items = soup.select('div[data-track-item]')

            log_message(task_id, f"   找到 {len(items)} 个商品元素")

            ad_rank, nat_rank = 0, 0

            for idx, item in enumerate(items):
                try:
                    full_text = " ".join(item.get_text().split())

                    link = "N/A"
                    title = "N/A"
                    review_link = ""

                    # 评论链接
                    for a in item.select('a[href*="review.rakuten.co.jp"]'):
                        href = a.get('href', '')
                        if 'review.rakuten.co.jp/item/' in href:
                            review_link = href
                            break

                    # 商品链接
                    for a in item.select('a[href*="item.rakuten.co.jp"]'):
                        href = a.get('href', '')
                        if 'item.rakuten.co.jp' in href:
                            link = href
                            title = a.get_text(strip=True) or "N/A"
                            break

                    if link == "N/A":
                        title_tag = item.select_one('.title a') or item.select_one('h2 a')
                        if title_tag:
                            link = title_tag.get('href', 'N/A')
                            title = title_tag.get_text(strip=True) or "N/A"

                    # 从属性获取链接
                    if link == "N/A" or 'redirect' in link:
                        item_id_attr = item.get('data-track-ratid') or item.get('data-item-id') or ""
                        if item_id_attr:
                            parts = item_id_attr.split(':')
                            if len(parts) >= 2:
                                shop_name = parts[0]
                                product_id = parts[1]
                                link = f"https://item.rakuten.co.jp/{shop_name}/{product_id}/"

                    if review_link and ('redirect' in link or 'grp' in link):
                        link = f"{link}|||{review_link}"

                    shop = item.select_one('.merchant a').get_text(strip=True) if item.select_one('.merchant a') else "N/A"

                    # 主图
                    img = "N/A"
                    for img_el in item.select('img'):
                        img_url = img_el.get('src') or img_el.get('data-src') or img_el.get('data-lazy') or ""
                        if not img_url:
                            continue
                        if any(x in img_url.lower() for x in ['.svg', '/assets/', '/resources/', 'logo', 'icon', 'badge', '39shop']):
                            continue
                        if any(x in img_url.lower() for x in ['.jpg', '.jpeg', '.png', '.webp', '.gif']):
                            img = img_url
                            break
                    if img and img != "N/A":
                        img = re.sub(r'\?.*$', '', img)
                        img = img.replace('_ex=80x80', '').replace('_ex=128x128', '')

                    # 价格
                    price = extract_price(full_text)
                    if price == "0":
                        pt = item.select_one('[class*="price"]')
                        if pt:
                            price = extract_price(pt.get_text())

                    # 评论
                    rev_cnt, rev_score = 0, "0.0"
                    score_match = re.search(r'[★☆評価]\s*(\d\.\d{1,2})', full_text)
                    if not score_match:
                        score_match = re.search(r'(\d\.\d{1,2})\s*[\(（]\d+[件\)）]', full_text)
                    if score_match:
                        score_val = float(score_match.group(1))
                        if 1.0 <= score_val <= 5.0:
                            rev_score = score_match.group(1)

                    cnt_match = re.search(r'[\(（](\d{1,6})[件]?[\)）]', full_text)
                    if cnt_match:
                        rev_cnt = int(cnt_match.group(1))

                    # 排名
                    is_ad = "[PR]" in title or item.select_one('.marker-pr') or "r.rakuten.co.jp" in link
                    if is_ad:
                        ad_rank += 1
                    else:
                        nat_rank += 1

                    sat = f"商品数{total_txt}件中, {('PR' if is_ad else '自然')}{ad_rank if is_ad else nat_rank}位"

                    raw_data.append({
                        "品牌": shop,
                        "标题": title,
                        "url": link,
                        "主图": img,
                        "饱和度": sat,
                        "评论数": rev_cnt,
                        "评分": rev_score,
                        "价格": int(price),
                        "上线时长": "...",
                        "预估月销": "",
                        "商品详细参数": "...",
                        "核心卖点分析": "...",
                        "评论出现优点": "...",
                        "客诉点": "...",
                        "有无视频": "...",
                        "大类排名": "",
                        "小类排名": "",
                        "备注": "",
                        "特征1": "", "特征2": "", "特征3": "", "特征4": "", "特征5": "",
                        "特征6": "", "特征7": "", "特征8": "", "特征9": ""
                    })
                except Exception as e:
                    log_message(task_id, f"   ⚠️ 商品 {idx+1} 解析失败: {e}")
                    continue

            task['progress'] = int((page / pages) * 20)
            log_message(task_id, f"   ✅ 本页获取 {len(items)} 条数据")
            time.sleep(1.5)

        if not raw_data:
            task['status'] = 'error'
            task['message'] = '未找到任何商品'
            log_message(task_id, "❌ 未找到任何商品")
            return

        # ==================== Step 2: 深度抓取 ====================
        df = pd.DataFrame(raw_data)
        limit = min(len(df), deep_limit) if deep_limit else len(df)

        log_message(task_id, f"🕵️ 开始深度抓取 (共 {limit} 条)...")

        for i in range(limit):
            row = df.iloc[i]
            is_pr = "PR广告" if "grp" in str(row['url']) else "普通"
            task['message'] = f'深度分析 {i+1}/{limit}: {row["标题"][:25]}...'
            task['progress'] = 20 + int((i / limit) * 60)
            log_message(task_id, f"   [{i+1}/{limit}] 分析中... [{is_pr}]")

            # 调用原始爬虫的完整深度抓取函数
            details = get_product_details(row['url'], row['评论数'])

            # 检查解析结果
            if details['上线时长'] == 'URL解析失败':
                log_message(task_id, f"      ⚠️ URL解析失败")

            df.at[i, '上线时长'] = details['上线时长']
            df.at[i, '商品详细参数'] = details['商品详细参数']
            df.at[i, '核心卖点分析'] = details['核心卖点分析']
            df.at[i, '评论出现优点'] = details['评论出现优点']
            df.at[i, '客诉点'] = details['客诉点']
            df.at[i, '有无视频'] = details['有无视频']
            df.at[i, '大类排名'] = details['大类排名']
            df.at[i, '小类排名'] = details['小类排名']
            df.at[i, '备注'] = details['备注']

            # 产品特征
            if details.get('产品特征'):
                for j, feat in enumerate(details['产品特征'][:9], 1):
                    df.at[i, f'特征{j}'] = feat
                log_message(task_id, f"      [AI] 提取到 {len(details['产品特征'])} 个特征")

            # 更新主图
            if details.get('主图') and (df.at[i, '主图'] == "N/A" or not df.at[i, '主图']):
                df.at[i, '主图'] = details['主图']

            # 补充评论数和评分
            if details.get('评论数_补充') is not None:
                try:
                    if int(df.at[i, '评论数']) == 0:
                        df.at[i, '评论数'] = details['评论数_补充']
                        log_message(task_id, f"      → 补充评论数: {details['评论数_补充']}")
                except:
                    df.at[i, '评论数'] = details['评论数_补充']

            if details.get('评分_补充') is not None:
                try:
                    if float(df.at[i, '评分']) == 0.0:
                        df.at[i, '评分'] = details['评分_补充']
                        log_message(task_id, f"      → 补充评分: {details['评分_补充']}")
                except:
                    df.at[i, '评分'] = details['评分_补充']

            time.sleep(random.uniform(1.2, 2.5))

        log_message(task_id, "✅ 深度抓取完成")

        # ==================== Step 3: 整理数据 ====================
        task['message'] = '正在整理数据格式...'
        task['progress'] = 85
        log_message(task_id, "📊 整理数据格式...")

        # 清理URL
        df['url'] = df['url'].apply(lambda x: x.split('|||')[0] if '|||' in str(x) else x)

        # 复制用于翻译
        df_jp = df.copy()

        # 添加关键词列
        df_jp['キーワード'] = keyword
        df_jp['キャッチコピー'] = df_jp['标题']
        df_jp['品牌（店铺名）'] = df_jp['品牌']
        df_jp['市场饱和度'] = df_jp['饱和度']
        df_jp['review数量'] = df_jp['评论数']
        df_jp['review评分'] = df_jp['评分']
        df_jp['价格（JPY)'] = df_jp['价格']
        df_jp['上线时长（月）'] = df_jp['上线时长']
        df_jp['月销售额'] = ""

        # 策略分析列
        df_jp['预估售价'] = ""
        df_jp['供应商是否可以开发票，采购多少个可以开'] = ""
        df_jp['平均毛利率'] = ""
        df_jp['促销频率'] = ""
        df_jp['可优化方向'] = ""
        df_jp['优先级'] = ""

        # 最终列顺序
        final_cols = [
            "品牌（店铺名）", "キーワード", "キャッチコピー", "url", "主图",
            "市场饱和度", "小类排名", "大类排名", "review数量", "review评分",
            "价格（JPY)", "上线时长（月）", "预估月销", "月销售额",
            "特征1", "特征2", "特征3", "特征4", "特征5", "特征6", "特征7", "特征8", "特征9",
            "核心卖点分析", "评论出现优点", "客诉点", "有无视频", "备注",
            "预估售价", "供应商是否可以开发票，采购多少个可以开", "平均毛利率", "促销频率", "可优化方向", "优先级"
        ]

        # 确保列存在
        for col in final_cols:
            if col not in df_jp.columns:
                df_jp[col] = ""

        df_jp = df_jp.reindex(columns=final_cols)

        # 中文翻译版
        df_cn = df_jp.copy()
        if enable_translate and ENABLE_TRANSLATION:
            task['message'] = '正在翻译中文版...'
            task['progress'] = 90
            log_message(task_id, "🇨🇳 正在翻译中文版...")

            cols_to_translate = ["品牌（店铺名）", "キャッチコピー", "市场饱和度",
                                 "核心卖点分析", "评论出现优点", "客诉点",
                                 "特征1", "特征2", "特征3", "特征4", "特征5",
                                 "特征6", "特征7", "特征8", "特征9"]

            for idx, col in enumerate(cols_to_translate):
                if col in df_cn.columns:
                    log_message(task_id, f"   翻译列: {col}")
                    try:
                        df_cn[col] = df_cn[col].apply(safe_translate)
                    except Exception as e:
                        log_message(task_id, f"   ⚠️ 翻译 {col} 失败: {e}")

            # 翻译排名
            def translate_rank(rank_text):
                if not rank_text or not isinstance(rank_text, str):
                    return rank_text
                match = re.match(r'(.+?)\s*第(\d+)', rank_text)
                if match:
                    cat_name_jp = match.group(1)
                    rank_num = match.group(2)
                    cat_name_cn = safe_translate(cat_name_jp)
                    return f"{cat_name_cn} 第{rank_num}"
                return rank_text

            df_cn['大类排名'] = df_cn['大类排名'].apply(translate_rank)
            df_cn['小类排名'] = df_cn['小类排名'].apply(translate_rank)
            log_message(task_id, "   ✅ 翻译完成")

        # 保存结果
        task['progress'] = 95
        task['message'] = '保存结果...'

        # 转换为结果列表
        results = df.to_dict('records')
        task['results'] = results
        task['dataframe_jp'] = df_jp
        task['dataframe_cn'] = df_cn
        task['keyword'] = keyword
        task['status'] = 'completed'
        task['progress'] = 100
        task['message'] = f'完成！共抓取 {len(results)} 条商品数据'
        log_message(task_id, f"🎉 完成！共抓取 {len(results)} 条商品数据")

    except Exception as e:
        task['status'] = 'error'
        task['message'] = f'爬虫错误: {str(e)}'
        log_message(task_id, f"❌ 爬虫错误: {str(e)}")
        import traceback
        log_message(task_id, traceback.format_exc())


# ==================== API 路由 ====================

@app.route('/')
def index():
    """首页"""
    return render_template('index_full.html')


@app.route('/api/scrape', methods=['POST'])
def start_scrape():
    """启动爬虫任务"""
    data = request.json
    keyword = data.get('keyword', '').strip()
    pages = int(data.get('pages', 1))
    deep_limit = int(data.get('deep_limit', 5))
    enable_ai = data.get('enable_ai', True)
    enable_translate = data.get('enable_translate', True)

    if not keyword:
        return jsonify({'error': '请输入关键词'}), 400

    task_id = str(uuid.uuid4())
    tasks[task_id] = {
        'id': task_id,
        'keyword': keyword,
        'status': 'pending',
        'progress': 0,
        'message': '任务已创建',
        'results': [],
        'dataframe_jp': None,
        'dataframe_cn': None,
        'created_at': datetime.now().isoformat()
    }
    task_logs[task_id] = []

    # 启动后台线程
    thread = threading.Thread(
        target=run_scraper_full,
        args=(task_id, keyword, pages, deep_limit, enable_ai, enable_translate)
    )
    thread.daemon = True
    thread.start()

    return jsonify({'task_id': task_id})


@app.route('/api/task/<task_id>')
def get_task_status(task_id):
    """获取任务状态"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    return jsonify({
        'id': task['id'],
        'keyword': task.get('keyword', ''),
        'status': task['status'],
        'progress': task['progress'],
        'message': task['message'],
        'results': task.get('results', []),
        'result_count': len(task.get('results', []))
    })


@app.route('/api/logs/<task_id>')
def get_task_logs(task_id):
    """获取任务日志"""
    logs = task_logs.get(task_id, [])
    return jsonify({'logs': logs})


@app.route('/api/download/<task_id>/<lang>')
def download_excel(task_id, lang='jp'):
    """下载Excel文件"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    if task['status'] != 'completed':
        return jsonify({'error': '任务未完成'}), 400

    df = task.get('dataframe_cn' if lang == 'cn' else 'dataframe_jp')
    if df is None:
        return jsonify({'error': '数据不存在'}), 400

    # 创建Excel
    output = io.BytesIO()

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "爬虫结果"

        # 样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 分组信息
        groups = [
            ("基本信息", 5),
            ("售卖信息", 9),
            ("产品特征与用户体验", 14),
            ("策略分析与复盘", 6)
        ]

        # 写入分组标题行
        col_idx = 1
        for group_name, group_cols in groups:
            cell = ws.cell(row=1, column=col_idx, value=group_name)
            cell.font = Font(bold=True)
            cell.fill = group_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if group_cols > 1:
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + group_cols - 1)
            col_idx += group_cols

        # 写入列头
        for c_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

        # 写入数据
        for r_idx, row in enumerate(df.values, 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border

        # 列宽
        col_widths = {
            "品牌（店铺名）": 18, "キーワード": 15, "キャッチコピー": 35, "url": 25, "主图": 20,
            "市场饱和度": 25, "小类排名": 15, "大类排名": 15, "review数量": 10, "review评分": 10,
            "价格（JPY)": 12, "上线时长（月）": 14, "预估月销": 12, "月销售额": 12,
            "特征1": 18, "特征2": 18, "特征3": 18, "特征4": 18, "特征5": 18,
            "特征6": 18, "特征7": 18, "特征8": 18, "特征9": 18,
            "核心卖点分析": 35, "评论出现优点": 35, "客诉点": 35, "有无视频": 10, "备注": 20,
            "预估售价": 12, "供应商是否可以开发票，采购多少个可以开": 25,
            "平均毛利率": 12, "促销频率": 12, "可优化方向": 20, "优先级": 10
        }

        for c_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(c_idx)
            ws.column_dimensions[col_letter].width = col_widths.get(col_name, 15)

        # 行高
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        for row_idx in range(3, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 60

        # 冻结
        ws.freeze_panes = 'A3'

        wb.save(output)
        output.seek(0)

    except Exception as e:
        return jsonify({'error': f'Excel生成失败: {str(e)}'}), 500

    lang_suffix = "CN" if lang == 'cn' else "JP"
    filename = f"rakuten_{task['keyword']}_{lang_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/config')
def get_config():
    """获取当前配置"""
    return jsonify({
        'proxy_port': PROXY_PORT,
        'ai_enabled': USE_AI_FEATURES,
        'translation_enabled': ENABLE_TRANSLATION
    })


if __name__ == '__main__':
    os.makedirs('templates', exist_ok=True)
    print("=" * 60)
    print("🚀 乐天商品爬虫 Web 应用 - 完整版")
    print(f"📡 代理端口: {PROXY_PORT}")
    print(f"🤖 AI特征提取: {'启用' if USE_AI_FEATURES else '禁用'}")
    print(f"🇨🇳 翻译功能: {'启用' if ENABLE_TRANSLATION else '禁用'}")
    print("=" * 60)
    print("访问 http://localhost:5000 开始使用")
    print("=" * 60)
    app.run(host='0.0.0.0', port=5000, debug=True, threaded=True)
