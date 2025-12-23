import pandas as pd
from curl_cffi import requests
from bs4 import BeautifulSoup
import time
import random
import re
import os
import io
from deep_translator import GoogleTranslator
from urllib.parse import unquote
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openai import OpenAI

# ==================== ⚙️ 配置区域 ====================
KEYWORD = "コカコーラ"
PAGES_TO_SCRAPE = 1
# ⚠️⚠️⚠️ 请务必确认您的代理端口
PROXY_PORT = 7897
# 深度抓取限制 (None=抓取本页全部, 5=只测前5个)
DEEP_SCRAPE_LIMIT = 5

# ==================== 🤖 AI配置 ====================
# API Key 从环境变量获取，或直接填写
AI_API_KEY = "sk-w8xPygiSwsGuoN5yeddcH373PofeRw5Vxcb3yhPmD92ga2UL"
AI_BASE_URL = "https://api.probex.top/v1"
AI_MODEL = "Qwen3-VL-235B-A22B-Instruct"
# 是否启用AI提取特征 (如果API Key为空则自动禁用)
USE_AI_FEATURES = bool(AI_API_KEY)

# 初始化AI客户端（走代理访问API）
import httpx

ai_client = None
if AI_API_KEY:
    try:
        # 创建走代理的 http 客户端
        http_client = httpx.Client(
            timeout=60.0,
            proxy=f"http://127.0.0.1:{PROXY_PORT}"
        )
        ai_client = OpenAI(
            api_key=AI_API_KEY,
            base_url=AI_BASE_URL,
            http_client=http_client
        )
        print(f"✅ AI特征提取已启用 (模型: {AI_MODEL})")
    except Exception as e:
        print(f"⚠️ AI初始化失败: {e}，将使用关键词匹配")
        USE_AI_FEATURES = False
else:
    print("⚠️ 未设置 AI_API_KEY，使用关键词匹配提取特征")
# 列表页抓取限制 (None=抓取本页全部, 5=只取前5个)
LIST_SCRAPE_LIMIT = 5
# 是否启用翻译 (True=翻译, False=不翻译，避免崩溃)
ENABLE_TRANSLATION = True
# 是否下载图片并嵌入Excel (True=下载嵌入, False=只保存链接)
DOWNLOAD_IMAGES = True
# 图片保存文件夹
IMAGE_FOLDER = "product_images"
# ====================================================

proxies = {
    "http": f"http://127.0.0.1:{PROXY_PORT}",
    "https": f"http://127.0.0.1:{PROXY_PORT}"
}


def safe_translate(text):
    """安全翻译函数，带重试和延迟"""
    if not text or str(text) == "nan" or text == "N/A" or text == "" or text == "...":
        return text
    try:
        text = str(text).strip()
        if len(text) < 2:  # 太短不翻译
            return text
        # 限制长度防止翻译API报错
        text = text[:800]
        time.sleep(0.3)  # 添加延迟避免请求过快
        result = GoogleTranslator(source='ja', target='zh-CN').translate(text)
        return result if result else text
    except Exception as e:
        # 翻译失败返回原文
        return text


def extract_features_from_images(image_urls, max_images=3):
    """
    【视觉AI特征提取】使用Qwen VL从商品图片中提取产品特征
    """
    global ai_client
    if not ai_client or not USE_AI_FEATURES or not image_urls:
        return []

    features = []
    import base64
    from io import BytesIO

    for img_url in image_urls[:max_images]:
        try:
            # 下载图片
            resp = requests.get(img_url, impersonate="chrome120", proxies=proxies, timeout=10)
            if resp.status_code != 200:
                continue

            # 转为base64
            img_b64 = base64.b64encode(resp.content).decode('utf-8')

            # 调用视觉AI
            response = ai_client.chat.completions.create(
                model=AI_MODEL,
                temperature=0,
                max_tokens=200,
                messages=[
                    {"role": "system", "content": "你是产品特征识别助手。请识别图片中展示的产品特征，用简短中文描述（2-6字），逗号分隔。"},
                    {"role": "user", "content": [
                        {"type": "text", "text": "请从这张商品图片中提取产品特征，如材质、颜色、结构、功能等。只输出特征词，用逗号分隔。"},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}}
                    ]}
                ]
            )
            result = response.choices[0].message.content.strip()
            # 解析
            img_features = [f.strip() for f in result.split(',') if f.strip() and len(f.strip()) <= 15]
            for f in img_features:
                if f not in features:
                    features.append(f)

            time.sleep(0.5)  # 避免请求过快
        except Exception as e:
            print(f"      [AI图片] 分析失败: {e}")
            continue

    return features[:9]


def extract_product_features_ai(title, description, specs, image_urls=None):
    """
    【真AI特征提取】使用Qwen大模型从商品信息中提取9个核心产品特征
    支持文本+图片联合分析
    """
    global ai_client
    if not ai_client or not USE_AI_FEATURES:
        return None

    features = []

    # 1. 从图片提取特征（如果有图片URL）
    if image_urls:
        img_features = extract_features_from_images(image_urls, max_images=2)
        features.extend(img_features)
        if img_features:
            print(f"      [AI图片] 提取到特征: {img_features[:5]}")

    # 2. 从文本提取特征
    prompt = f"""请从以下日本乐天商品信息中提取最多9个核心产品卖点特征。

要求：
1. 每个特征用简短中文描述（2-8个字）
2. 提取产品的核心卖点，如：材质、功能、设计、适用场景等
3. 参考示例：开放式收纳、防锈涂层、坚固耐用、带轮可移动、高度可调、组装简单、大容量、多层设计、节省空间

商品标题：{title[:200]}
商品描述：{description[:500]}
商品参数：{specs[:400]}

请直接返回9个特征，用逗号分隔，不要其他解释。"""

    try:
        response = ai_client.chat.completions.create(
            model=AI_MODEL,
            temperature=0,
            max_tokens=200,
            messages=[
                {"role": "system", "content": "你是一个产品特征提取助手，只输出简短的中文特征词，用逗号分隔。"},
                {"role": "user", "content": prompt}
            ]
        )
        result = response.choices[0].message.content.strip()
        # 解析结果
        text_features = [f.strip() for f in result.split(',') if f.strip() and len(f.strip()) <= 15]
        for f in text_features:
            if f not in features:
                features.append(f)
    except Exception as e:
        print(f"      [AI文本] 调用失败: {e}")

    return features[:9] if features else None


def extract_product_features_keywords(title, description, specs):
    """
    【关键词匹配】备用方案：从商品信息中用关键词匹配提取特征
    """
    features = []

    # 合并所有文本
    all_text = f"{title} {description} {specs}".lower()

    # 特征关键词库 (日文关键词 -> 中文特征)
    feature_keywords = {
        # 收纳方式
        "オープン|開放|見せる収納": "开放式收纳",
        "扉付き|引き出し|隠す収納": "封闭式收纳",
        # 材质
        "スチール|鉄|金属|メタル": "钢铁材质",
        "木製|ウッド|天然木": "木质材质",
        "プラスチック|樹脂": "塑料材质",
        "ステンレス": "不锈钢材质",
        # 表面处理
        "防錆|サビ防止|錆びにくい|粉体塗装": "防锈涂层",
        "防水|耐水": "防水处理",
        "クロムメッキ|メッキ": "电镀处理",
        # 结构强度
        "頑丈|丈夫|耐荷重|強い|堅牢": "坚固耐用",
        "軽量|軽い": "轻便",
        # 移动性
        "キャスター|車輪|移動": "带轮可移动",
        "アジャスター|固定脚|安定": "固定脚稳定",
        # 可调节
        "高さ調節|調整可能|可動": "高度可调",
        "棚板.*調節|段階調整": "层板可调",
        # 透气性
        "メッシュ|網|通気|通風": "网状透气",
        "ワイヤー": "钢丝网结构",
        # 组装
        "簡単組立|工具不要|ワンタッチ": "组装简单",
        "組み立て式": "需简单组装",
        # 外观
        "おしゃれ|スタイリッシュ|モダン": "外观时尚",
        "省スペース|スリム|コンパクト": "节省空间",
        # 用途
        "キッチン|台所": "适合厨房",
        "洗面所|浴室|バス": "适合浴室",
        "押入れ|クローゼット": "适合衣柜",
        "リビング|居間": "适合客厅",
        # 其他特点
        "大容量|たっぷり収納": "大容量",
        "多機能|多用途": "多功能",
        "伸縮|拡張": "可伸缩",
        # 新增：层数/段数
        r"\d+段|\d+層": "多层设计",
        "4段|四段": "4层设计",
        "5段|五段": "5层设计",
        # 新增：品牌/品质
        "ルミナス|luminous": "知名品牌",
        "日本製|国産": "日本制造",
        "業務用|プロ": "专业级",
        # 新增：尺寸
        r"幅\d+|横幅": "宽度适中",
        r"奥行\d+": "深度适中",
        r"高さ\d+cm": "高度适中",
    }

    # 匹配特征
    for keywords, feature in feature_keywords.items():
        if re.search(keywords, all_text, re.I):
            if feature not in features:
                features.append(feature)

    # 如果特征不足，尝试从参数中提取更多信息
    if len(features) < 5:
        # 尺寸信息
        size_match = re.search(r'(幅|奥行|高さ)[^\d]*(\d+)', all_text)
        if size_match and "尺寸规格" not in features:
            features.append("尺寸规格明确")

        # 颜色
        colors = re.findall(r'(ブラック|ホワイト|シルバー|ブラウン|ナチュラル|黒|白|銀)', all_text)
        if colors and "多色可选" not in features:
            features.append("多色可选" if len(set(colors)) > 1 else f"颜色简约")

    return features[:9]


def extract_product_features(title, description, specs, image_urls=None):
    """
    【智能特征提取】优先使用AI（支持图片分析），失败则用关键词匹配
    """
    # 优先使用AI（包含图片分析）
    if USE_AI_FEATURES:
        ai_features = extract_product_features_ai(title, description, specs, image_urls)
        if ai_features:
            return ai_features

    # 备用：关键词匹配
    return extract_product_features_keywords(title, description, specs)


def download_image(img_url, save_path):
    """
    下载图片并保存到本地
    返回: 本地文件路径 或 None
    """
    if not img_url or img_url == "N/A":
        return None
    try:
        resp = requests.get(img_url, impersonate="chrome120", proxies=proxies, timeout=10)
        if resp.status_code == 200:
            # 确保目录存在
            os.makedirs(os.path.dirname(save_path), exist_ok=True)

            # 用PIL处理图片（统一转为PNG，调整大小）
            img_data = io.BytesIO(resp.content)
            img = PILImage.open(img_data)

            # 转换为RGB（处理RGBA或其他模式）
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')

            # 调整大小（Excel显示用，宽度80px）
            max_width = 80
            ratio = max_width / img.width
            new_height = int(img.height * ratio)
            img = img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)

            # 保存为JPEG
            save_path = save_path.rsplit('.', 1)[0] + '.jpg'
            img.save(save_path, 'JPEG', quality=85)
            return save_path
    except Exception as e:
        pass
    return None


def resolve_pr_link(url: str) -> str:
    """
    【修复】强力解析 PR 广告链接 (支持 Meta Refresh 和 JS 跳转)
    """
    if not url: return ""
    if "item.rakuten.co.jp" in url and "redirect" not in url: return url

    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        resp = requests.get(
            url,
            impersonate="chrome120",
            proxies=proxies,
            timeout=15,
            allow_redirects=True,
            headers=headers
        )

        if "item.rakuten.co.jp" in resp.url: return resp.url

        soup = BeautifulSoup(resp.text, 'html.parser')

        # 方法1: Meta Refresh
        meta = soup.find("meta", attrs={"http-equiv": re.compile("refresh", re.I)})
        if meta:
            content = meta.get("content", "")
            m = re.search(r'url=([^"]+)', content, re.I)
            if m: return resolve_pr_link(m.group(1))

        # 方法2: JS window.location.replace
        js = re.search(r'window\.location\.replace\("([^"]+)"\)', resp.text)
        if js: return resolve_pr_link(js.group(1))

        # 方法3: JS window.location.href 或 location.href
        js2 = re.search(r'(?:window\.)?location\.href\s*=\s*["\']([^"\']+)["\']', resp.text)
        if js2: return resolve_pr_link(js2.group(1))

        # 方法4: 在页面内找 item.rakuten.co.jp 链接
        lk = re.search(r'href="([^"]*item\.rakuten\.co\.jp[^"]*)"', resp.text)
        if lk: return lk.group(1)

        # 方法5: 从 JSON 数据中提取（乐天广告经常用JSON）
        json_url = re.search(r'"(?:url|link|href)"\s*:\s*"(https?://item\.rakuten\.co\.jp[^"]+)"', resp.text)
        if json_url: return json_url.group(1)

        # 方法6: 从任意位置提取 item.rakuten.co.jp 链接
        any_item = re.search(r'(https?://item\.rakuten\.co\.jp/[^/]+/[^/\s"\'<>]+)', resp.text)
        if any_item: return any_item.group(1)

        return url
    except Exception as e:
        print(f"      [DEBUG] resolve_pr_link 异常: {e}")
        return url


def extract_dynamic_specs(soup):
    """
    【核心功能】动态抓取详情页的所有表格参数
    """
    specs_list = []
    tables = soup.find_all('table')
    for table in tables:
        rows = table.find_all('tr')
        for row in rows:
            cols = row.find_all(['th', 'td'])
            if len(cols) == 2:
                key = cols[0].get_text(strip=True)
                val = cols[1].get_text(strip=True)
                if len(key) < 30 and len(val) < 200 and val:
                    if any(x in key for x in ["配送", "支払", "送料", "カード", "あす楽"]): continue
                    specs_list.append(f"【{key}】: {val}")

    if not specs_list:
        lis = soup.select('.item_desc li')
        for li in lis:
            t = li.get_text(strip=True)
            if ":" in t or "：" in t: specs_list.append(t)

    if specs_list:
        return "\n".join(specs_list[:20])
    return "未抓取到参数"


# ==================== ✨ 新增功能函数区域 ✨ ====================

def check_has_video(soup, raw_html):
    """检测页面是否包含视频"""
    # 1. 检查 video 标签
    if soup.find("video"): return "有"
    # 2. 检查乐天专用播放器 class
    if soup.select(".rakutenVideoPlayer"): return "有"
    # 3. 检查 HTML 源码中的视频特征
    if "rakuten.co.jp/rms/mall/image/video" in raw_html: return "有"
    # 4. 检查 iframe (Youtube/Vimeo)
    iframes = soup.find_all("iframe")
    for iframe in iframes:
        src = iframe.get("src", "")
        if "youtube" in src or "vimeo" in src: return "有"
    return "无"


def get_categories_from_page(soup, raw_html=""):
    """
    从商品详情页自动提取类目ID和名称
    返回: [{"id": "566374", "name": "スチールラック"}, ...]
    """
    categories = []
    page_text = raw_html if raw_html else str(soup)

    def add_category(cat_id, name):
        """添加类目（去重）"""
        if cat_id and len(cat_id) >= 5 and not any(c['id'] == cat_id for c in categories):
            categories.append({"id": cat_id, "name": name or f"类目{cat_id}"})

    try:
        # 方法1: 从面包屑导航提取 (/category/)
        for bc in soup.select('a[href*="/category/"]'):
            href = bc.get('href', '')
            name = bc.get_text(strip=True)
            match = re.search(r'/category/(\d+)/', href)
            if match and len(name) < 30:
                add_category(match.group(1), name)

        # 方法2: 从 genre 链接提取 (/genre/) - 总是执行
        for gl in soup.select('a[href*="/genre/"]'):
            href = gl.get('href', '')
            name = gl.get_text(strip=True)
            match = re.search(r'/genre/(\d+)', href)
            if match and len(name) < 30:
                add_category(match.group(1), name)

        # 方法3: 从 ranking 链接提取 - 总是执行
        for link in soup.select('a[href*="ranking.rakuten.co.jp"]'):
            href = link.get('href', '')
            match = re.search(r'ranking\.rakuten\.co\.jp/\w+/(\d+)', href)
            if match:
                name = link.get_text(strip=True)
                add_category(match.group(1), name)

        # 方法4: 从页面内嵌的JSON数据提取 (多种格式) - 总是执行
        # 匹配 genreId, categoryId, genre_id, category_id 等
        genre_matches = re.findall(r'["\'](?:genre|category)[_]?[Ii]d["\']\s*:\s*["\']?(\d{5,})', page_text)
        for gid in genre_matches[:10]:
            add_category(gid, None)

        # 方法5: 从 JSON-LD 或 script 标签提取 - 总是执行
        # 匹配 "genreId":"123456" 或 genreId=123456 或 "genreId":123456
        all_genres = re.findall(r'genre[Ii]d["\'\s:=]+["\']?(\d{5,})', page_text)
        for gid in all_genres[:10]:
            add_category(gid, None)

        # 方法6: 从 URL 参数提取 - 总是执行
        url_genres = re.findall(r'[?&]genre[_]?id=(\d{5,})', page_text, re.I)
        for gid in url_genres[:5]:
            add_category(gid, None)

        # 方法7: 从常见的乐天类目格式提取
        # 匹配 "l-id=xxx_yyy_zzzzz" 格式中的数字ID
        lid_matches = re.findall(r'l-id=[^"\'&]*?(\d{6})', page_text)
        for gid in lid_matches[:5]:
            add_category(gid, None)

        # 调试输出
        if categories:
            print(f"      [类目DEBUG] 提取到 {len(categories)} 个类目: {[c['id'] for c in categories[:5]]}")
        else:
            print(f"      [类目DEBUG] ⚠️ 未能提取到类目信息")

    except Exception as e:
        print(f"      [类目DEBUG] 提取异常: {e}")

    return categories


# ==================== ⚙️ 排行榜配置 ====================
RANKING_SEARCH_PAGES = 3  # 搜索深度


# =======================================================

def search_product_in_ranking(shop_id, item_id, category_id, rank_type="daily"):
    """
    通用排名查找函数 (HTML切割法 - 终极精准版)
    逻辑：
    1. 在源码中搜索“X位”这个排名标记 (如 'alt="1位"' 或 'alt="81位"')。
    2. 找到后，把这个标记 **之前** 的所有 HTML 代码全部切除！
    3. 这样就彻底删除了顶部的“浏览记录”和“广告”干扰。
    4. 在剩下的纯净代码里数数，商品排第几，就是第几名。
    """
    base_url = f"https://ranking.rakuten.co.jp/{rank_type}/{category_id}/"

    # 目标：店铺名/商品ID 组合 (小写)
    target_shop = shop_id.lower()
    target_item = item_id.lower()
    target_full = f"{target_shop}/{target_item}"  # 完整匹配

    try:
        for page in range(1, RANKING_SEARCH_PAGES + 1):
            url = base_url
            if page > 1: url += f"?p={page}"

            print(f"        🔍 [正在检查] p{page}: {url} ...", end="", flush=True)

            resp = requests.get(url, impersonate="chrome120", proxies=proxies, timeout=15)
            if resp.status_code != 200:
                print(" ❌ HTTP错误")
                return None, None

            # 转小写用于搜索
            full_html = resp.text.lower()

            # 1. 过滤无效页面
            if "ページが表示できません" in full_html:
                print(" ⚠️ 无效榜单，跳过")
                return None, None

            # 2. 检查重定向
            if "総合ランキング" in full_html and rank_type not in full_html:
                # 简单检查标题，防止误判
                soup_check = BeautifulSoup(resp.text, 'html.parser')
                h1 = soup_check.select_one('h1')
                if h1 and "総合ランキング" in h1.get_text():
                    print(" ⚠️ 转跳综合榜，跳过")
                    return None, None

            # --- ✂️ 核心逻辑：寻找切入点 (Anchor) ---

            # 我们需要找到榜单开始的地方。特征通常是 "1位", "46位", "81位" 等图片或文字
            # 匹配 alt="数字位" 或者 >数字位<
            # 我们寻找页面上出现的第一个“排名数字”

            start_anchor_rank = 1  # 默认从1名开始
            slice_index = 0  # 默认不切割

            # 正则搜索第一个出现的排名标记
            # 模式：alt="123位" 或 class="rank">123位
            match = re.search(r'(?:alt="|class="[^"]*rank[^"]*".*?>|>\s*)(\d{1,3})\s*位', full_html)

            if match:
                start_anchor_rank = int(match.group(1))
                slice_index = match.end()
            else:
                # 如果没找到标记，根据页码估算起始值 (45 或 80)
                # print(" [未定位] 使用默认起始", end="")
                if page > 1:
                    # 尝试判断是45还是80模式，默认先给80 (乐天现在很多是80)
                    start_anchor_rank = (page - 1) * 80 + 1

            # --- ✂️ 执行切割！ ---
            # 只保留从排名开始位置之后的 HTML
            # 这样顶部的 History、Ads 全部被扔掉了
            clean_html = full_html[slice_index:]

            # --- 🔮 在纯净 HTML 中查找商品 ---
            # 提取所有商品链接：店铺名/商品ID 组合
            raw_matches = re.findall(r'item\.rakuten\.co\.jp/([^/]+)/([^/"\?]+)', clean_html)

            # ⚠️ 关键修复：去重！每个商品只计数一次（保持顺序）
            seen = set()
            found_ids = []  # 存储 "店铺名/商品ID" 组合
            for shop_found, item_found in raw_matches:
                full_id = f"{shop_found}/{item_found}"
                if full_id not in seen:
                    seen.add(full_id)
                    found_ids.append(full_id)

            if target_full in found_ids:
                final_rank = None
                idx_in_list = found_ids.index(target_full)
                print(f" [idx={idx_in_list},共{len(found_ids)}个]", end="")

                # 方法1: 从 shopUnit[X] JSON数据中提取排名
                shop_unit_matches = list(re.finditer(r'shopUnit\[(\d+)\]', full_html))
                if shop_unit_matches:
                    # 找到目标商品链接的位置（包含店铺名）
                    target_pattern = rf'item\.rakuten\.co\.jp/{re.escape(target_shop)}/{re.escape(target_item)}'
                    target_match = re.search(target_pattern, full_html, re.IGNORECASE)

                    if target_match:
                        target_pos = target_match.start()
                        # 找最接近目标（且在目标之前）的 shopUnit
                        for m in reversed(shop_unit_matches):
                            if m.start() < target_pos:
                                final_rank = int(m.group(1))
                                print(f" [JSON:shopUnit[{final_rank}]]", end="")
                                break

                # 方法2: 计数法（使用切割后的列表）
                if final_rank is None:
                    idx = found_ids.index(target_full)
                    final_rank = start_anchor_rank + idx

                    # ⚠️ 修正：检查目标商品是否在切割后列表中"提前"出现
                    # 如果目标商品在原始HTML中的第一次出现位置，比切割后列表中的第一次出现位置更靠前
                    # 说明有商品被漏掉了，需要+1
                    if page == 1 and start_anchor_rank == 1:
                        # 找目标在切割前HTML中的位置（使用完整的店铺名/商品ID）
                        target_pattern_full = rf'item\.rakuten\.co\.jp/{re.escape(target_shop)}/{re.escape(target_item)}'
                        target_in_full = re.search(target_pattern_full, full_html)
                        target_in_clean = re.search(target_pattern_full, clean_html)

                        if target_in_full and target_in_clean:
                            # 用更宽松的正则，匹配所有可能的商品链接格式
                            # 包括 item.rakuten.co.jp/店铺/商品 和编码格式
                            pattern = r'item\.rakuten\.co\.jp/([^/"\s]+)/([^/"\s\?]+)'

                            # 计算目标在切割前HTML中是第几个出现的商品
                            matches_full = re.findall(pattern, full_html[:target_in_full.start()])
                            matches_clean = re.findall(pattern, clean_html[:target_in_clean.start()])

                            # 用 (店铺, 商品) 组合来去重
                            unique_full = len(set(matches_full))
                            unique_clean = len(set(matches_clean))

                            # 直接用 full 计数作为排名（更准确）
                            final_rank = unique_full + 1
                            print(f" [full={unique_full}+1={final_rank}]", end="")
                        else:
                            print(f" [计数={final_rank}]", end="")
                    else:
                        print(f" [计数={final_rank}]", end="")

                # 获取类名用于返回
                cat_name = ""
                soup = BeautifulSoup(resp.text, 'html.parser')
                title_el = soup.select_one('h1') or soup.select_one('.title')
                if title_el:
                    cat_name = title_el.get_text(strip=True).replace("ランキング", "").replace("デイリー", "").strip()

                print(f" ✅ 排名: {final_rank}名")
                return final_rank, cat_name
            else:
                print(" ❌ 不在此页榜单中")

    except Exception as e:
        print(f" ⚠️ 报错: {e}")
        pass

    return None, None


def get_category_ranking(shop_id, item_id, page_categories):
    """
    获取排名主逻辑 (带调试信息)
    """
    result = {"大类排名": "", "小类排名": ""}
    if not shop_id or not item_id: return result

    found_ranks = []

    # 去重类目ID
    unique_cats = {c['id']: c for c in page_categories}.values()

    print(f"      🚀 准备扫描 {len(unique_cats)} 个类目...")

    # 只查日榜 (daily) - 您目前的配置
    check_types = ["daily"]

    for cat in unique_cats:
        for r_type in check_types:
            rank, cat_name = search_product_in_ranking(shop_id, item_id, cat["id"], r_type)

            if rank:
                display_name = cat_name if cat_name else cat["name"]
                found_ranks.append({
                    "rank": rank,
                    "name": display_name,
                    "id_len": len(cat["id"]),
                })

    # 总结结果
    if found_ranks:
        # 排序：优先取ID长的(小类)，其次看排名靠前的
        found_ranks.sort(key=lambda x: (-x["id_len"], x["rank"]))

        best = found_ranks[0]
        result["小类排名"] = f"{best['name']} 第{best['rank']}"
        print(f"      🎉 最终锁定: 小类[{result['小类排名']}]", end="")

        if len(found_ranks) > 1:
            # 找一个名字不一样的大类
            for r in found_ranks[1:]:
                if r["name"] != best["name"]:
                    result["大类排名"] = f"{r['name']} 第{r['rank']}"
                    print(f" / 大类[{result['大类排名']}]")
                    break

            # 如果没找到不一样名字的，填第二个
            if not result["大类排名"]:
                sec = found_ranks[1]
                result["大类排名"] = f"{sec['name']} 第{sec['rank']}"
                print(f" / 大类[{result['大类排名']}]")
        else:
            print("")  # 换行
    else:
        print("      🤷‍♂️ 未在任何榜单找到排名")

    return result


def extract_ranking_info(soup, raw_html, shop_id="", item_id=""):
    """从详情页提取大类排名和小类排名，支持排行榜反查"""
    result = {"大类排名": "", "小类排名": ""}

    try:
        page_text = soup.get_text()

        # 方法1: 匹配排名徽章文本 "XXXランキング X位"
        rank_patterns = re.findall(r'([^\s]{2,20}ランキング)[^\d]*(\d{1,4})位', page_text)

        for category, rank in rank_patterns:
            rank_text = f"{category} {rank}位"
            if any(kw in category for kw in ["デイリー", "週間", "リアルタイム"]):
                if not result["小类排名"]:
                    result["小类排名"] = rank_text
            else:
                if not result["大类排名"]:
                    result["大类排名"] = rank_text

        # 方法2: 从HTML属性/class中查找排名元素
        rank_elements = soup.select('[class*="rank"]') or soup.select('[class*="Rank"]')
        for el in rank_elements:
            text = el.get_text(strip=True)
            match = re.search(r'(\d{1,4})位', text)
            if match:
                rank_num = match.group(1)
                if not result["小类排名"]:
                    result["小类排名"] = f"排名 {rank_num}位"
                    break

        # 方法3: 从乐天排行榜页面反查 (获取 "类目名 第X" 格式) - 优先级最高！
        if shop_id and item_id:
            # 自动从页面提取类目ID
            page_categories = get_categories_from_page(soup, raw_html)
            if page_categories:
                print(f"      → 排行榜反查中 (发现{len(page_categories)}个类目)...")
                ranking_data = get_category_ranking(shop_id, item_id, page_categories)
                # 反查结果优先覆盖（更准确）
                if ranking_data["大类排名"]:
                    result["大类排名"] = ranking_data["大类排名"]
                if ranking_data["小类排名"]:
                    result["小类排名"] = ranking_data["小类排名"]

    except Exception as e:
        print(f"      [排名] 提取失败: {e}")

    return result


def extract_selling_points(soup):
    """提取核心卖点 (Catch Copy + 商品简介)"""
    points = []
    # 1. Catch Copy (通常在标题上方的红色/加粗文字)
    catch_copy = soup.select_one('.catch_copy')
    if catch_copy:
        points.append(catch_copy.get_text(strip=True))

    # 2. 商品描述的前几项 (通常是核心卖点)
    lis = soup.select('.item_desc li')
    for li in lis[:3]:
        points.append(li.get_text(strip=True))

    if not points:
        # 备选：找主要描述段落
        desc = soup.select_one('.item_desc')
        if desc: points.append(desc.get_text(strip=True)[:100])

    return "\n".join(points) if points else "无明显卖点"


def analyze_selling_points_ai(raw_points, title=""):
    """
    【AI核心卖点分析】使用AI对提取的日文卖点进行分析总结
    输出简洁的中文卖点描述
    """
    global ai_client
    if not ai_client or not USE_AI_FEATURES:
        return raw_points

    if not raw_points or raw_points == "无明显卖点":
        return raw_points

    prompt = f"""请分析以下日本乐天商品的核心卖点，用简洁的中文总结出3-5个核心卖点。

商品标题：{title[:100]}
原始卖点描述：
{raw_points[:800]}

要求：
1. 每个卖点用简短中文描述（10-20字）
2. 提取最有价值的卖点信息（品牌优势、产品特点、使用场景等）
3. 用换行分隔每个卖点

请直接输出卖点列表，不要其他解释。"""

    try:
        response = ai_client.chat.completions.create(
            model=AI_MODEL,
            temperature=0,
            max_tokens=300,
            messages=[
                {"role": "system", "content": "你是产品卖点分析师，擅长从日文商品描述中提取核心卖点并翻译成简洁中文。"},
                {"role": "user", "content": prompt}
            ]
        )
        result = response.choices[0].message.content.strip()
        if result:
            print(f"      [AI卖点] 分析完成")
            return result
    except Exception as e:
        print(f"      [AI卖点] 分析失败: {e}")

    return raw_points  # 失败时返回原文


def summarize_reviews_ai(pros_list, cons_list):
    """
    【AI评论总结】将评论提炼成简短关键词
    输出格式：
    - 评论出现优点：组装简单（简短关键词）
    - 客诉点：1.架子变形 2.包装破损（问题列表）
    """
    global ai_client
    res = {
        "评论出现优点": "暂无评论",
        "客诉点": "无明显差评"
    }

    if not ai_client or not USE_AI_FEATURES:
        # 无AI时，简单截取
        if pros_list:
            res["评论出现优点"] = pros_list[0][:50] if pros_list else "暂无评论"
        if cons_list:
            res["客诉点"] = "\n".join([f"{i + 1}.{c[:30]}" for i, c in enumerate(cons_list[:3])])
        return res

    # 准备评论文本 - 采样更多评论以获得更准确的分析
    # 从好评中均匀采样30条
    import random
    if len(pros_list) > 30:
        step = len(pros_list) // 30
        sampled_pros = [pros_list[i] for i in range(0, len(pros_list), step)][:30]
    else:
        sampled_pros = pros_list
    pros_text = "\n".join(sampled_pros)

    # 差评全部使用（通常数量不多）
    cons_text = "\n".join(cons_list[:15]) if cons_list else "无"

    prompt = f"""请分析以下日本乐天商品的用户评论，提炼该商品独特的评论特点。

【好评内容】（共{len(pros_list)}条，以下为采样{len(sampled_pros)}条）:
{pros_text[:3000]}

【差评/客诉】（{len(cons_list)}条）:
{cons_text[:1500]}

请按以下格式输出（必须用中文）：

评论出现优点：[列出3-5个该商品被用户提到最多的具体优点，要体现差异化]
- 避免泛泛的词如"性价比高、质量好"
- 优先提取具体特点，如"防锈好、承重强、棚板厚实、颜色漂亮、配件齐全、客服好"
- 用空格分隔

客诉点：[列出1-3个用户抱怨的具体问题，如"1.螺丝生锈 2.棚板有刮痕 3.说明书难懂"，没有差评写"无明显差评"]

注意：提取评论中出现频率高的具体描述，避免笼统词汇。"""

    try:
        response = ai_client.chat.completions.create(
            model=AI_MODEL,
            temperature=0,
            max_tokens=200,
            messages=[
                {"role": "system", "content": "你是评论分析师，擅长从日文评论中提炼关键信息，输出简洁中文。"},
                {"role": "user", "content": prompt}
            ]
        )
        result = response.choices[0].message.content.strip()

        # 解析AI输出
        if "评论出现优点" in result or "优点" in result:
            # 提取优点
            import re
            pros_match = re.search(r'(?:评论出现)?优点[：:]\s*(.+?)(?:\n|客诉|$)', result)
            if pros_match:
                res["评论出现优点"] = pros_match.group(1).strip()[:50]

            # 提取客诉点
            cons_match = re.search(r'客诉点[：:]\s*(.+?)$', result, re.DOTALL)
            if cons_match:
                res["客诉点"] = cons_match.group(1).strip()[:100]

        print(f"      [AI评论] 总结完成: 优点={res['评论出现优点'][:15]}...")
        return res

    except Exception as e:
        print(f"      [AI评论] 总结失败: {e}")
        # 失败时用简单方式
        if pros_list:
            res["评论出现优点"] = pros_list[0][:50]
        if cons_list:
            res["客诉点"] = "\n".join([f"{i + 1}.{c[:30]}" for i, c in enumerate(cons_list[:3])])
        return res


def analyze_reviews(review_url, max_pages=10):
    """
    【新增】进入评论页，分析优缺点和客诉
    自动检测总页数并抓取所有评论
    """
    res = {
        "评论出现优点": "暂无评论",
        "客诉点": "无明显差评"
    }
    if not review_url or "http" not in review_url: return res

    try:
        pros = []  # 4-5星
        cons = []  # 1-2星 (客诉)

        # 构建基础URL（去掉末尾的排序参数）
        # 评论URL格式: https://review.rakuten.co.jp/item/1/306224_10008717/1.1/
        # 分页格式: https://review.rakuten.co.jp/item/1/306224_10008717/?p=2
        base_url = re.sub(r'/\d+\.\d+/?$', '/', review_url)
        print(f"      [评论DEBUG] base_url: {base_url}")

        # 先获取第一页，检测总页数（用基础URL，不带排序参数）
        first_resp = requests.get(base_url, impersonate="chrome120", proxies=proxies, timeout=10)

        # 从页面中提取总页数（多种方式）
        # 方式1: 匹配 ?p=X 的最大值
        page_nums = re.findall(r'\?p=(\d+)', first_resp.text)
        # 方式2: 匹配 "X页中" 或 "全X页"
        if not page_nums:
            total_match = re.search(r'全(\d+)ページ|(\d+)ページ中', first_resp.text)
            if total_match:
                page_nums = [total_match.group(1) or total_match.group(2)]
        # 方式3: 从评论总数计算（每页约30条）
        if not page_nums:
            total_count_match = re.search(r'"reviewCount"\s*:\s*"?(\d+)"?', first_resp.text)
            if total_count_match:
                total_count = int(total_count_match.group(1))
                calculated_pages = (total_count + 29) // 30  # 每页约30条，向上取整
                page_nums = [str(calculated_pages)]
                print(f"      [评论DEBUG] 从评论总数{total_count}计算得{calculated_pages}页")

        if page_nums:
            total_pages = min(int(max(page_nums, key=int)), max_pages)  # 最多抓max_pages页
        else:
            total_pages = 1
        print(f"      [评论] 检测到 {total_pages} 页评论，开始抓取...")

        # 抓取所有页评论
        total_reviews = 0
        for page in range(1, total_pages + 1):
            page_url = f"{base_url}?p={page}" if page > 1 else review_url
            try:
                resp = requests.get(page_url, impersonate="chrome120", proxies=proxies, timeout=10)
                page_text = resp.text

                # 乐天2024新版评论格式：评论文本在 </div></div><div class="expand-link 之前
                review_pattern = r'>([^<]{20,500})</div></div><div class="expand-link'
                matches = re.findall(review_pattern, page_text)

                if not matches:
                    break  # 没有更多评论了

                # 提取评论文本
                for text in matches:
                    text = text.strip().replace('\n', ' ')[:200]
                    if len(text) > 15 and not text.startswith('<'):
                        # 简单判断好评/差评（根据关键词）
                        negative_words = ['残念', '悪い', 'がっかり', '最悪', '壊れ', '傷', 'ダメ', '不良', '欠け', '凹', '錆']
                        if any(w in text for w in negative_words):
                            cons.append(text)
                        else:
                            pros.append(text)
                        total_reviews += 1

                time.sleep(0.3)  # 避免请求过快
            except Exception as e:
                print(f"      [评论] 第{page}页抓取失败: {e}")
                break

        print(f"      [评论DEBUG] 共抓取 {total_pages} 页, 找到 {total_reviews} 条评论 (好评{len(pros)}/差评{len(cons)})")

        # 用AI总结评论（简短关键词格式）
        if pros or cons:
            res = summarize_reviews_ai(pros, cons)

        return res
    except Exception as e:
        print(f"      [评论分析] 错误: {e}")
        return res


# ==============================================================


def get_product_details(item_url, review_count_check=1):
    """
    深度抓取主逻辑：整合了 上线时间 + 参数 + 卖点 + 视频 + 评论分析
    """
    result = {
        "上线时长": "等待抓取...",
        "商品详细参数": "等待抓取...",
        "核心卖点分析": "...",
        "评论出现优点": "...",
        "客诉点": "...",
        "有无视频": "无",
        "大类排名": "",
        "小类排名": "",
        "主图": "",  # 从详情页获取高清主图
        "备注": None,
        "评论数_补充": None,  # 从详情页补充的评论数
        "评分_补充": None,  # 从详情页补充的评分
        "产品特征": [],  # AI提取的产品特征列表
    }

    try:
        # 检查是否有备用评论链接 (格式: 原始链接|||评论链接)
        review_backup = ""
        if "|||" in item_url:
            parts = item_url.split("|||")
            item_url = parts[0]
            review_backup = parts[1] if len(parts) > 1 else ""

        # 1. 解析真实 URL
        real_url = resolve_pr_link(item_url)
        real_url = unquote(real_url)
        print(f"      [DEBUG] 解析后URL: {real_url[:80]}...")

        m = re.search(r'item\.rakuten\.co\.jp/([^/?#]+)/([^/?#]+)', real_url)
        if not m:
            # 尝试备用解析：从 product.rakuten.co.jp 格式
            m = re.search(r'product\.rakuten\.co\.jp/product/-/([^/?#]+)', real_url)
            if m:
                # 这是产品聚合页，尝试直接访问获取商品链接
                print(f"      [DEBUG] 检测到产品聚合页，尝试访问...")
                try:
                    prod_resp = requests.get(real_url, impersonate="chrome120", proxies=proxies, timeout=15)
                    m2 = re.search(r'item\.rakuten\.co\.jp/([^/?#"]+)/([^/?#"]+)', prod_resp.text)
                    if m2:
                        m = m2
                except:
                    pass

        # 🔥 备用方案：从评论页获取商品链接
        if not m and review_backup:
            print(f"      [DEBUG] 尝试从评论页获取商品链接: {review_backup[:60]}...")
            try:
                review_resp = requests.get(review_backup, impersonate="chrome120", proxies=proxies, timeout=15)
                # 从评论页HTML中提取商品链接
                m = re.search(r'item\.rakuten\.co\.jp/([^/?#"\']+)/([^/?#"\']+)', review_resp.text)
                if m:
                    print(f"      [DEBUG] ✅ 从评论页成功提取商品链接!")
            except Exception as e:
                print(f"      [DEBUG] 评论页访问失败: {e}")

        if not m:
            result["上线时长"] = "URL解析失败"
            result["备注"] = f"无法解析URL: {real_url[:50]}"
            return result

        shop_id, item_id = m.group(1), m.group(2)
        item_page_url = f"https://item.rakuten.co.jp/{shop_id}/{item_id}/"
        print(f"      [DEBUG] 商品页URL: {item_page_url}")

        # 2. 请求商品详情页
        item_resp = requests.get(item_page_url, impersonate="chrome120", proxies=proxies, timeout=15)
        if item_resp.status_code != 200: return result

        item_soup = BeautifulSoup(item_resp.text, 'html.parser')

        # --- A. 抓取参数 ---
        result["商品详细参数"] = extract_dynamic_specs(item_soup)

        # --- ✨ 新增: 从详情页补充评论数和评分 ---
        # 直接从HTML中用更精确的正则匹配乐天评分格式
        # 典型格式: "4.63(1,526件)" 或在JSON中 "reviewAverage":"4.63","reviewCount":1526

        # 方法1: 从JSON-LD或脚本中提取
        json_score = re.search(r'"ratingValue"\s*:\s*"?(\d\.\d+)"?', item_resp.text)
        json_count = re.search(r'"reviewCount"\s*:\s*"?(\d+)"?', item_resp.text)
        if json_score and json_count:
            result["评分_补充"] = json_score.group(1)
            result["评论数_补充"] = int(json_count.group(1))

        # 方法2: 从页面文本匹配 4.63(1,526件)
        if result["评分_补充"] is None:
            review_match = re.search(r'(\d\.\d{1,2})\s*[\(（]([0-9,]+)\s*件[\)）]', item_resp.text)
            if review_match:
                result["评分_补充"] = review_match.group(1)
                result["评论数_补充"] = int(review_match.group(2).replace(',', ''))

        # 方法3: 分开匹配
        if result["评分_补充"] is None:
            score_m = re.search(r'>(\d\.\d{1,2})<', item_resp.text)  # 标签内的评分
            cnt_m = re.search(r'[\(（]([0-9,]+)\s*件[\)）]', item_resp.text)
            if score_m and cnt_m:
                score_val = float(score_m.group(1))
                if 1.0 <= score_val <= 5.0:
                    result["评分_补充"] = score_m.group(1)
                    result["评论数_补充"] = int(cnt_m.group(1).replace(',', ''))

        # --- ✨ 新增: 抓取高清主图 ---
        main_img = ""
        # 尝试多种选择器获取主图
        img_el = (item_soup.select_one('meta[property="og:image"]') or
                  item_soup.select_one('.rakutenLimitedId_ImageMain1-3 img') or
                  item_soup.select_one('.image-main img') or
                  item_soup.select_one('[class*="mainImage"] img') or
                  item_soup.select_one('.item-image img'))
        if img_el:
            main_img = img_el.get('content') or img_el.get('src') or ""
        # 备用：从页面源码正则提取
        if not main_img:
            img_match = re.search(r'"image"\s*:\s*"(https://[^"]+\.(?:jpg|jpeg|png|webp))"', item_resp.text, re.I)
            if img_match:
                main_img = img_match.group(1)
        result["主图"] = main_img

        # --- ✨ 新增: 抓取核心卖点 ---
        raw_selling_points = extract_selling_points(item_soup)

        # --- ✨ 新增: AI提取产品特征 ---
        # 获取标题
        title_tag = item_soup.find('title')
        title = title_tag.get_text(strip=True) if title_tag else ""

        # 用AI分析核心卖点（翻译+总结）
        result["核心卖点分析"] = analyze_selling_points_ai(raw_selling_points, title)
        # 获取描述
        desc_tag = item_soup.select_one('.item_desc') or item_soup.select_one('[class*="description"]')
        description = desc_tag.get_text(strip=True)[:500] if desc_tag else ""

        # 提取详情页中的产品展示图片（用于AI视觉分析）
        detail_images = []
        if USE_AI_FEATURES:
            # 🔥 优先使用商品主图（最准确）
            if main_img and 'http' in main_img:
                detail_images.append(main_img)

            # 从商品描述区域提取图片（严格过滤，排除广告区域）
            # 只选择商品描述区域的图片，不要用 .rakutenLimitedId（太宽泛会包含广告）
            desc_imgs = item_soup.select('.item_desc img, .item-image img, [class*="itemImage"] img')
            for img in desc_imgs[:8]:
                src = img.get('src') or img.get('data-src') or ''
                if src and 'http' in src and any(ext in src.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                    # 严格过滤：排除广告、促销、季节活动图
                    skip_words = ['icon', 'logo', 'banner', 'campaign', 'sale', 'point', 'review',
                                  'cart', 'btn', 'button', 'arrow', 'star', 'rank', 'pr_', 'ad_',
                                  'winter', 'summer', 'spring', 'autumn', 'season', 'event', 'special',
                                  'entry', 'coupon', 'deal', 'stamp', 'subscription', '定期購入',
                                  'snowman', 'tea', 'coffee', 'hand', 'warm', 'tokushu', 'guide']
                    if any(w in src.lower() for w in skip_words):
                        continue
                    # 只保留真正的商品图片
                    if ('image.rakuten' in src or 'shop.r10s' in src or 'thumbnail' in src) and shop_id in src:
                        if src not in detail_images:
                            detail_images.append(src)

            # 限制最多5张图片
            detail_images = detail_images[:5]
            if detail_images:
                print(f"      [AI] 发现 {len(detail_images)} 张商品图待分析")

        # 合并所有文本信息用于特征提取（包含核心卖点）
        full_description = f"{description}\n{result['核心卖点分析']}"

        # 提取特征（支持图片分析）
        result["产品特征"] = extract_product_features(title, full_description, result["商品详细参数"], detail_images)

        # --- ✨ 新增: 抓取有无视频 ---
        result["有无视频"] = check_has_video(item_soup, item_resp.text)

        # --- ✨ 新增: 抓取排名信息 (支持排行榜反查) ---
        print(f"      [排名DEBUG] 正在查询: shop={shop_id}, item={item_id}")
        ranking_info = extract_ranking_info(item_soup, item_resp.text, shop_id, item_id)
        result["大类排名"] = ranking_info["大类排名"]
        result["小类排名"] = ranking_info["小类排名"]
        # 保存翻译后的版本（如果有）
        result["大类排名_CN"] = ranking_info.get("大类排名_CN", "")
        result["小类排名_CN"] = ranking_info.get("小类排名_CN", "")

        # --- B. 抓取上线时间 与 评论分析 ---
        # 如果列表页评论数是0，但详情页可能有评论，用详情页的数据重新判断
        actual_review_count = review_count_check
        if int(review_count_check) == 0 and result.get("评论数_补充"):
            actual_review_count = result["评论数_补充"]
            print(f"      [DEBUG] 列表页评论数0，但详情页有评论: {actual_review_count}")

        if int(actual_review_count) == 0:
            result["上线时长"] = "暂无评论(新品)"
            result["评论出现优点"] = "无评论"
            result["客诉点"] = "无评论"
        else:
            # 在详情页里找评论链接
            review_match = re.search(r'href="(https://review\.rakuten\.co\.jp/item/1/[^"]+)"', item_resp.text)
            if review_match:
                review_url = review_match.group(1)

                # --- ✨ 新增: 深度分析评论内容 (优点/客诉) ---
                # 注意：这里我们单独请求一次评论页，既为了拿时间，也为了拿内容
                # 为了代码复用，我们可以在 analyze_reviews 里拿内容，在这里单独拿时间，或者合并
                # 这里为了稳健，分别处理

                # 1. 获取优缺点
                print(f"      [DEBUG] 正在分析评论: {review_url}")
                review_data = analyze_reviews(review_url)
                result["评论出现优点"] = review_data["评论出现优点"]
                result["客诉点"] = review_data["客诉点"]
                print(f"      [DEBUG] 评论结果: 优点={review_data['评论出现优点'][:30]}...")

                time.sleep(0.5)

                # 2. 获取上线时间 - 直接访问按时间从旧到新排序的评论页
                # 评论URL格式: https://review.rakuten.co.jp/item/1/306224_10008717/1.1/
                # 需要去掉末尾的 /x.x/ 部分
                base_review_url = review_url.split('?')[0]
                base_review_url = re.sub(r'/\d+\.\d+/?$', '', base_review_url)

                # 先访问第一页获取总评论数和计算最后一页
                rev_resp = requests.get(base_review_url, impersonate="chrome120", proxies=proxies, timeout=15)

                # 从JSON中提取总评论数 "nr_max_review":169
                total_match = re.search(r'"nr_max_review"\s*:\s*(\d+)', rev_resp.text)
                if total_match:
                    total_reviews = int(total_match.group(1))
                    # 每页约15条，计算最后一页（但不超过实际页数）
                    # 先用较小的估算，每页20条
                    max_page = max(1, (total_reviews + 19) // 20)
                    # 访问最后一页
                    last_page_url = f"{base_review_url}?p={max_page}"
                    print(f"      [DEBUG] 评论总数: {total_reviews}, 访问最后一页: {last_page_url}")
                    rev_resp = requests.get(last_page_url, impersonate="chrome120", proxies=proxies, timeout=15)

                    # 如果最后一页没有内容，逐页往前找
                    decoded_text = rev_resp.text.replace('\\u002F', '/')
                    if '"orderDate"' not in decoded_text and max_page > 1:
                        for try_page in range(max_page - 1, 0, -1):
                            try_url = f"{base_review_url}?p={try_page}"
                            rev_resp = requests.get(try_url, impersonate="chrome120", proxies=proxies, timeout=15)
                            decoded_text = rev_resp.text.replace('\\u002F', '/')
                            if '"orderDate"' in decoded_text:
                                print(f"      [DEBUG] 实际最后一页: p={try_page}")
                                break
                else:
                    print(f"      [DEBUG] 无法获取评论总数，使用第一页")

                # 匹配评论日期 - 从JSON中提取 orderDate
                # 格式: "orderDate":"2023\u002F04\u002F10" 或 "orderDate":"2023/04/10"
                # 先解码 \u002F 为 /
                decoded_text = rev_resp.text.replace('\\u002F', '/')

                # 从JSON中提取 orderDate
                dates = re.findall(r'"orderDate"\s*:\s*"(20\d{2})/(\d{1,2})/(\d{1,2})"', decoded_text)
                print(f"      [DEBUG] JSON orderDate匹配: {dates[:5]}")

                # 备用：从页面文本中匹配 注文日:2023/02/10
                if not dates:
                    dates = re.findall(r'注文日.{0,3}(20\d{2})/(\d{1,2})/(\d{1,2})', rev_resp.text)
                    print(f"      [DEBUG] 注文日格式匹配: {dates[:5]}")

                # 过滤无效日期（月份1-12，日期1-31，且年份>=2010）
                valid_dates = []
                for y, m, d in dates:
                    try:
                        year = int(y)
                        month = int(m)
                        day = int(d)
                        if 2010 <= year <= 2025 and 1 <= month <= 12 and 1 <= day <= 31:
                            valid_dates.append(f"{y}-{str(m).zfill(2)}-{str(d).zfill(2)}")
                    except:
                        pass

                print(f"      [DEBUG] 有效日期: {sorted(valid_dates)[:5]}")

                if valid_dates:
                    valid_dates.sort()
                    result["上线时长"] = valid_dates[0]  # 最早的评论日期
                else:
                    result["上线时长"] = "日期提取失败"
            else:
                result["上线时长"] = "无评论链接"

        return result

    except Exception as e:
        print(f"详情页错误: {e}")
        return result


def extract_price(text):
    if not text: return "0"
    text = str(text).replace(',', '').replace(' ', '')
    match = re.search(r'(\d+)円', text)
    if match: return match.group(1)
    match = re.search(r'[¥￥](\d+)', text)
    if match: return match.group(1)
    match = re.search(r'(\d{3,})', text)
    if match: return match.group(1)
    return "0"


def run_spider():
    print(f"🚀 启动【完全体整合版】爬虫 | 关键词: {KEYWORD}")
    print(f"📡 代理端口: {PROXY_PORT} (请确保正确)")
    print("-" * 60)

    base_url = "https://search.rakuten.co.jp/search/mall/{}/"
    raw_data = []

    # --- Step 1: 列表抓取 ---
    for page in range(1, PAGES_TO_SCRAPE + 1):
        url = base_url.format(KEYWORD) + f"?p={page}"
        print(f"📡 [Step 1] 抓取第 {page} 页列表...")

        try:
            res = requests.get(url, impersonate="chrome120", timeout=30, proxies=proxies)
            soup = BeautifulSoup(res.text, 'html.parser')

            # 市场饱和度 - 提取总商品数（格式：394,408件）
            total_txt = "未知"
            page_text = soup.get_text()

            # 方法1: 匹配 "検索結果1～45件（394,408件）" 格式
            count_match = re.search(r'[\(（]([\d,]+)件[\)）]', page_text)
            if count_match and len(count_match.group(1).replace(',', '')) >= 3:
                total_txt = count_match.group(1)

            # 方法2: 匹配 "394,408件中" 或 "394408 件"
            if total_txt == "未知":
                count_match = re.search(r'([\d,]{5,})件', page_text)
                if count_match:
                    total_txt = count_match.group(1)

            # 方法3: 从 span.count 等元素提取
            if total_txt == "未知":
                count_el = soup.select_one('.search-count') or soup.select_one('[class*="count"]')
                if count_el:
                    m = re.search(r'([\d,]+)', count_el.get_text())
                    if m and len(m.group(1).replace(',', '')) >= 3:
                        total_txt = m.group(1)

            print(f"   📊 搜索结果总数: {total_txt}件")

            items = soup.select('.searchresultitem')
            if not items: items = soup.select('div[data-track-item]')

            # DEBUG: 保存HTML用于分析
            with open("debug_list_page.html", "w", encoding="utf-8") as f:
                f.write(res.text)
            print(f"      [DEBUG] 已保存列表页HTML到 debug_list_page.html")
            print(f"      [DEBUG] 找到 {len(items)} 个商品元素 (选择器: .searchresultitem)")

            # 打印第一个item的所有a标签
            if items:
                first_item = items[0]
                all_links = first_item.select('a[href]')
                print(f"      [DEBUG] 第1个商品内有 {len(all_links)} 个链接")
                for i, a in enumerate(all_links[:5]):
                    print(f"        链接{i + 1}: {a.get('href', '')[:80]}")

            # 限制列表页抓取数量
            if LIST_SCRAPE_LIMIT:
                items = items[:LIST_SCRAPE_LIMIT]

            ad_rank, nat_rank = 0, 0

            for idx, item in enumerate(items):
                try:
                    full_text = " ".join(item.get_text().split())

                    # 多种方式尝试获取商品链接
                    link = "N/A"
                    title = "N/A"
                    review_link = ""  # 保存评论链接作为备用

                    # 🔥 方式1: 保存评论链接（用于备用解析）
                    for a in item.select('a[href*="review.rakuten.co.jp"]'):
                        href = a.get('href', '')
                        if 'review.rakuten.co.jp/item/' in href:
                            review_link = href
                            break

                    # 方式2: 直接找包含 item.rakuten.co.jp 的链接
                    for a in item.select('a[href*="item.rakuten.co.jp"]'):
                        href = a.get('href', '')
                        if 'item.rakuten.co.jp' in href:
                            link = href
                            title = a.get_text(strip=True) or "N/A"
                            break

                    # 方式3: .title a 或 h2 a（可能是redirect链接）
                    if link == "N/A":
                        title_tag = item.select_one('.title a') or item.select_one('h2 a')
                        if title_tag:
                            link = title_tag.get('href', 'N/A')
                            title = title_tag.get_text(strip=True) or "N/A"

                    # 方式4: 从 data-track-ratid 或 data-item-id 属性获取
                    if link == "N/A" or 'redirect' in link:
                        # 查找商品ID属性
                        item_id_attr = item.get('data-track-ratid') or item.get('data-item-id') or ""
                        if item_id_attr:
                            # 格式可能是 "店铺名:商品ID"
                            parts = item_id_attr.split(':')
                            if len(parts) >= 2:
                                shop_name = parts[0]
                                product_id = parts[1]
                                link = f"https://item.rakuten.co.jp/{shop_name}/{product_id}/"
                                print(f"      [DEBUG] 从属性构建链接: {link}")

                    # 🔥 方式5: 如果link仍是redirect，附加评论链接用于备用解析
                    if review_link and ('redirect' in link or 'grp' in link):
                        # 用特殊格式保存：原始链接|||评论链接
                        link = f"{link}|||{review_link}"

                    # 调试：显示原始链接（末尾部分，区分不同链接）
                    link_suffix = link[-50:] if len(link) > 50 else link
                    print(f"      [列表DEBUG] #{idx + 1} 链接末尾: ...{link_suffix}")
                    shop = item.select_one('.merchant a').get_text(strip=True) if item.select_one(
                        '.merchant a') else "N/A"
                    # 主图获取 - 多种选择器
                    img = "N/A"
                    # 优先获取商品图片（排除图标）
                    for img_el in item.select('img'):
                        img_url = img_el.get('src') or img_el.get('data-src') or img_el.get('data-lazy') or ""
                        # 过滤掉非商品图片
                        if not img_url:
                            continue
                        # 排除: svg图标、assets资源、logo、icon等
                        if any(x in img_url.lower() for x in
                               ['.svg', '/assets/', '/resources/', 'logo', 'icon', 'badge', '39shop']):
                            continue
                        # 必须是图片格式
                        if any(x in img_url.lower() for x in ['.jpg', '.jpeg', '.png', '.webp', '.gif']):
                            img = img_url
                            break
                    # 清理URL
                    if img and img != "N/A":
                        img = re.sub(r'\?.*$', '', img)  # 去掉URL参数
                        img = img.replace('_ex=80x80', '').replace('_ex=128x128', '')  # 去掉尺寸限制

                    # 价格提取
                    price = extract_price(full_text)
                    if price == "0":
                        pt = item.select_one('[class*="price"]')
                        if pt: price = extract_price(pt.get_text())

                    # Review - 更精确的提取
                    rev_cnt, rev_score = 0, "0.0"

                    # 评分：匹配 ★4.67 或 評価4.67 或 (4.67) 格式
                    score_match = re.search(r'[★☆評価]\s*(\d\.\d{1,2})', full_text)
                    if not score_match:
                        # 备选：匹配评论数前的小数，如 "4.67(169件)"
                        score_match = re.search(r'(\d\.\d{1,2})\s*[\(（]\d+[件\)）]', full_text)
                    if score_match:
                        score_val = float(score_match.group(1))
                        if 1.0 <= score_val <= 5.0:
                            rev_score = score_match.group(1)

                    # 评论数：匹配 (169件) 或 （169件）
                    cnt_match = re.search(r'[\(（](\d{1,6})[件]?[\)）]', full_text)
                    if cnt_match:
                        rev_cnt = int(cnt_match.group(1))

                    # 排名
                    is_ad = "[PR]" in title or item.select_one('.marker-pr') or "r.rakuten.co.jp" in link
                    if is_ad:
                        ad_rank += 1
                    else:
                        nat_rank += 1

                    sat = f"商品数{total_txt}件中, {('PR' if is_ad else '自然')}{ad_rank if is_ad else nat_rank}位"

                    # 这里先初始化所有的key，防止后面报错
                    raw_data.append({
                        "品牌": shop, "标题": title, "url": link, "主图": img, "饱和度": sat,
                        "评论数": rev_cnt, "评分": rev_score, "价格": int(price),
                        "上线时长": "...", "预估月销": "", "商品详细参数": "...",
                        "核心卖点分析": "...", "评论出现优点": "...", "客诉点": "...",
                        "有无视频": "...", "大类排名": "", "小类排名": "", "备注": "",
                        "特征1": "", "特征2": "", "特征3": "", "特征4": "", "特征5": "",
                        "特征6": "", "特征7": "", "特征8": "", "特征9": ""
                    })
                except:
                    continue
            print(f"   ✅ 本页获取 {len(items)} 条数据")
            time.sleep(2)
        except Exception as e:
            print(f"错误: {e}")

    # --- Step 2: 深度抓取 ---
    df = pd.DataFrame(raw_data)
    if df.empty: return
    limit = len(df) if DEEP_SCRAPE_LIMIT is None else min(len(df), DEEP_SCRAPE_LIMIT)

    print("-" * 60)
    print(f"🕵️ [Step 2] 深度抓取参数、时间、评论分析 (共 {limit} 条)...")

    for i in range(limit):
        row = df.iloc[i]
        is_pr = "PR广告" if "grp" in str(row['url']) else "普通"
        print(f"   [{i + 1}/{limit}] 分析中... [{is_pr}]")
        print(f"      [DEBUG] 原始URL: {str(row['url'])[:80]}...")

        # 调用核心深度抓取函数 (同时获取时间+参数+卖点+评论)
        details = get_product_details(row['url'], row['评论数'])

        # 检查是否解析成功
        if details['上线时长'] == 'URL解析失败':
            print(f"      ⚠️ URL解析失败，跳过此商品")

        df.at[i, '上线时长'] = details['上线时长']
        df.at[i, '商品详细参数'] = details['商品详细参数']
        # ✨ 填入新数据
        df.at[i, '核心卖点分析'] = details['核心卖点分析']
        df.at[i, '评论出现优点'] = details['评论出现优点']
        df.at[i, '客诉点'] = details['客诉点']
        df.at[i, '有无视频'] = details['有无视频']
        df.at[i, '大类排名'] = details['大类排名']
        df.at[i, '小类排名'] = details['小类排名']
        df.at[i, '备注'] = details['备注']

        # ✨ 保存AI提取的产品特征
        if details.get('产品特征'):
            for j, feat in enumerate(details['产品特征'][:9], 1):
                df.at[i, f'特征{j}'] = feat

        # 如果详情页获取到了更好的主图，更新它
        if details.get('主图') and (df.at[i, '主图'] == "N/A" or not df.at[i, '主图']):
            df.at[i, '主图'] = details['主图']

        # ✨ 用详情页数据补充缺失的评论数和评分
        if details.get('评论数_补充') is not None:
            current_cnt = df.at[i, '评论数']
            # 检查是否为0或缺失
            try:
                if int(current_cnt) == 0:
                    df.at[i, '评论数'] = details['评论数_补充']
                    print(f"      → 补充评论数: {details['评论数_补充']}")
            except:
                df.at[i, '评论数'] = details['评论数_补充']
                print(f"      → 补充评论数: {details['评论数_补充']}")
        if details.get('评分_补充') is not None:
            current_score = df.at[i, '评分']
            # 检查是否为0或"0.0"或缺失
            try:
                if float(current_score) == 0.0:
                    df.at[i, '评分'] = details['评分_补充']
                    print(f"      → 补充评分: {details['评分_补充']}")
            except:
                df.at[i, '评分'] = details['评分_补充']
                print(f"      → 补充评分: {details['评分_补充']}")

        # try:
        #     df.at[i, '预估月销'] = f"Reviews score: {int(df.at[i, '评论数'] * 1.5)}"
        # except:
        #     pass

        time.sleep(random.uniform(1.2, 2.5))

    print("\n✅ 抓取完成")

    # --- Step 3: 整理数据并保存 ---
    print("-" * 60)
    print("📊 [Step 3] 整理数据格式...")

    # 复制df用于处理
    df_jp = df.copy()

    # 最终列顺序 - 匹配目标表格格式
    # 基本信息
    df_jp['キーワード'] = KEYWORD
    df_jp['キャッチコピー'] = df_jp['标题']  # 使用标题作为Catch Copy
    df_jp['品牌（店铺名）'] = df_jp['品牌']

    # 售卖信息 - 添加缺失列
    df_jp['市场饱和度'] = df_jp['饱和度']
    df_jp['小类排名'] = df['小类排名']
    df_jp['大类排名'] = df['大类排名']
    df_jp['review数量'] = df_jp['评论数']
    df_jp['review评分'] = df_jp['评分']
    df_jp['价格（JPY)'] = df_jp['价格']
    df_jp['上线时长（月）'] = df_jp['上线时长']
    df_jp['月销售额'] = ""  # 需要计算

    # 产品特征 - 使用AI提取的特征，如果没有则用参数填充
    for i in range(1, 10):
        if f'特征{i}' not in df_jp.columns:
            df_jp[f'特征{i}'] = ""

    # 备选：如果AI特征为空，用商品详细参数填充
    for idx, row in df_jp.iterrows():
        # 检查是否有AI提取的特征
        has_ai_features = any(str(row.get(f'特征{i}', '')).strip() for i in range(1, 10))
        if not has_ai_features:
            params = str(row.get('商品详细参数', '')).split('\n')
            for i, param in enumerate(params[:9], 1):
                df_jp.at[idx, f'特征{i}'] = param

    # 策略分析与复盘 - 空列供手动填写
    df_jp['预估售价'] = ""
    df_jp['供应商是否可以开发票，采购多少个可以开'] = ""
    df_jp['平均毛利率'] = ""
    df_jp['促销频率'] = ""
    df_jp['可优化方向'] = ""
    df_jp['优先级'] = ""

    # 最终列顺序
    final_cols = [
        # 基本信息
        "品牌（店铺名）", "キーワード", "キャッチコピー", "url", "主图",
        # 售卖信息
        "市场饱和度", "小类排名", "大类排名", "review数量", "review评分",
        "价格（JPY)", "上线时长（月）", "预估月销", "月销售额",
        # 产品特征与用户体验
        "特征1", "特征2", "特征3", "特征4", "特征5", "特征6", "特征7", "特征8", "特征9",
        "核心卖点分析", "评论出现优点", "客诉点", "有无视频", "备注",
        # 策略分析与复盘
        "预估售价", "供应商是否可以开发票，采购多少个可以开", "平均毛利率", "促销频率", "可优化方向", "优先级"
    ]

    # 确保列都存在
    df_jp = df_jp.reindex(columns=final_cols)

    # 中文版
    df_cn = df_jp.copy()
    if ENABLE_TRANSLATION:
        print("🇨🇳 正在翻译中文版 (可能需要几分钟)...")
        cols_to_translate = ["品牌（店铺名）", "キャッチコピー", "市场饱和度",
                             "核心卖点分析", "评论出现优点", "客诉点",
                             "特征1", "特征2", "特征3", "特征4", "特征5",
                             "特征6", "特征7", "特征8", "特征9"]
        total_cols = len(cols_to_translate)
        for idx, col in enumerate(cols_to_translate, 1):
            if col in df_cn.columns:
                print(f"   [{idx}/{total_cols}] 正在翻译: {col} ...")
                try:
                    df_cn[col] = df_cn[col].apply(safe_translate)
                except Exception as e:
                    print(f"   ⚠️ 翻译 {col} 失败: {e}")

        # 翻译排名中的类目名
        def translate_rank(rank_text):
            if not rank_text or not isinstance(rank_text, str):
                return rank_text
            match = re.match(r'(.+?)\s*第(\d+)', rank_text)
            if match:
                cat_name_jp = match.group(1)
                rank_num = match.group(2)
                cat_name_cn = safe_translate(cat_name_jp)
                return f"{cat_name_cn} 第{rank_num}"
            return rank_text

        print("   [翻译排名...]")
        df_cn['大类排名'] = df_cn['大类排名'].apply(translate_rank)
        df_cn['小类排名'] = df_cn['小类排名'].apply(translate_rank)

        print("   ✅ 翻译完成")
    else:
        print("⏭️ 跳过翻译 (ENABLE_TRANSLATION=False)")

    def save_excel_with_format(df, filename, add_group_headers=True):
        """保存Excel并设置格式：分组标题、自动换行、调整列宽、调整行高"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active

            # 定义分组信息
            groups = [
                ("基本信息", 5),  # 5列
                ("售卖信息", 9),  # 9列
                ("产品特征与用户体验", 14),  # 14列
                ("策略分析与复盘", 6)  # 6列
            ]

            # 样式定义
            header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            start_row = 1
            if add_group_headers:
                # 写入分组标题行 (第1行)
                col_idx = 1
                for group_name, group_cols in groups:
                    cell = ws.cell(row=1, column=col_idx, value=group_name)
                    cell.font = Font(bold=True)
                    cell.fill = group_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    # 合并单元格
                    if group_cols > 1:
                        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1,
                                       end_column=col_idx + group_cols - 1)
                    col_idx += group_cols
                start_row = 2

            # 写入列头 (第2行)
            for c_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.border = thin_border

            # 找到主图列的索引
            img_col_idx = None
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == "主图":
                    img_col_idx = idx
                    break

            # 写入数据 (从第3行开始)
            for r_idx, row in enumerate(df.values, start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border

                    # 如果是主图列且启用了图片下载，嵌入图片
                    if DOWNLOAD_IMAGES and c_idx == img_col_idx and value and value != "N/A":
                        try:
                            # 生成唯一文件名
                            img_filename = f"{IMAGE_FOLDER}/img_{r_idx - start_row}.jpg"
                            local_path = download_image(value, img_filename)
                            if local_path and os.path.exists(local_path):
                                # 嵌入图片到单元格
                                img = XLImage(local_path)
                                img.width = 60
                                img.height = 60
                                # 定位到单元格
                                col_letter = get_column_letter(c_idx)
                                img.anchor = f"{col_letter}{r_idx}"
                                ws.add_image(img)
                                # 清空单元格文字（只保留图片）
                                cell.value = ""
                        except Exception as e:
                            pass  # 图片处理失败，保留URL

            # 设置列宽 - 使用 get_column_letter 函数
            col_widths = {
                "品牌（店铺名）": 18, "キーワード": 15, "キャッチコピー": 35, "url": 20, "主图": 15,
                "市场饱和度": 25, "小类排名": 12, "大类排名": 12, "review数量": 10, "review评分": 10,
                "价格（JPY)": 12, "上线时长（月）": 14, "预估月销": 12, "月销售额": 12,
                "特征1": 20, "特征2": 20, "特征3": 20, "特征4": 20, "特征5": 20,
                "特征6": 20, "特征7": 20, "特征8": 20, "特征9": 20,
                "核心卖点分析": 35, "评论出现优点": 40, "客诉点": 40, "有无视频": 10, "备注": 15,
                "预估售价": 12, "供应商是否可以开发票，采购多少个可以开": 20,
                "平均毛利率": 12, "促销频率": 12, "可优化方向": 20, "优先级": 10
            }

            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = col_widths.get(col_name, 15)

            # 设置行高
            ws.row_dimensions[1].height = 25  # 分组标题行
            ws.row_dimensions[2].height = 30  # 列头行
            for row_idx in range(3, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 80

            # 冻结前两行
            ws.freeze_panes = 'A3'

            wb.save(filename)
            print(f"🎉 已保存: {filename}")
        except Exception as e:
            print(f"Excel保存失败({e})，尝试CSV...")
            df.to_csv(filename.replace('.xlsx', '.csv'), index=False, encoding='utf-8-sig')

    if DOWNLOAD_IMAGES:
        print(f"🖼️ 启用图片下载，图片将保存到 {IMAGE_FOLDER}/ 文件夹并嵌入Excel")

    save_excel_with_format(df_jp, "rakuten_complete_JP.xlsx")
    save_excel_with_format(df_cn, "rakuten_complete_CN.xlsx")


if __name__ == "__main__":
    run_spider()